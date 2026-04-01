import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import numpy as np
import os
import threading
import traceback
import warnings # Added to silence the openpyxl warning
from openpyxl.styles import PatternFill, Font, Alignment

# --- Silence harmless openpyxl Data Validation warnings ---
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# ============================================================================
# RETRO WINDOWS 95/98 THEME CONSTANTS
# ============================================================================
WIN95_BG = "#D4D0C8"        
WIN95_BLUE = "#000080"      
WIN95_WHITE = "#FFFFFF"
WIN95_BLACK = "#000000"

FONT_MAIN = ('MS Sans Serif', 8)
FONT_BOLD = ('MS Sans Serif', 8, 'bold')

class RetroGSTApp:
    def __init__(self, root):
        self.root = root
        self.root.title("GST Report Auditor - System Utility")
        self.root.minsize(520, 480)
        self.root.geometry("520x480") 
        self.root.configure(bg=WIN95_BG)
        self.root.resizable(True, True)
        
        self.monthly_files = []
        self.ytd_file = ""
        self.output_file = ""
        self.sheet_vars = {} 

        self.setup_ui()

    def setup_ui(self):
        banner = tk.Frame(self.root, bg=WIN95_BLUE, height=40)
        banner.pack(fill="x")
        banner.pack_propagate(False)
        tk.Label(banner, text=" GST Utility", font=('Arial', 12, 'bold', 'italic'), fg=WIN95_WHITE, bg=WIN95_BLUE).pack(side="left", padx=10)

        main_frame = tk.Frame(self.root, bg=WIN95_BG, padx=15, pady=10)
        main_frame.pack(fill="both", expand=True)

        # 1. Data Sources
        group1 = tk.LabelFrame(main_frame, text=" 1. Data Sources ", font=FONT_MAIN, bg=WIN95_BG, relief=tk.GROOVE, bd=2, padx=10, pady=10)
        group1.pack(fill="x", pady=(0, 10))

        tk.Label(group1, text="Past Reports:", font=FONT_MAIN, bg=WIN95_BG).grid(row=0, column=0, sticky="w", pady=5)
        self.lbl_monthly = tk.Label(group1, text=" <No files selected>", font=FONT_MAIN, bg=WIN95_WHITE, relief=tk.SUNKEN, bd=2, anchor="w", width=40)
        self.lbl_monthly.grid(row=0, column=1, padx=10, pady=5)
        tk.Button(group1, text="Browse...", font=FONT_MAIN, bg=WIN95_BG, relief=tk.RAISED, bd=2, width=10, command=self.select_monthly_files).grid(row=0, column=2)

        tk.Label(group1, text="YTD Master:", font=FONT_MAIN, bg=WIN95_BG).grid(row=1, column=0, sticky="w", pady=5)
        self.lbl_ytd = tk.Label(group1, text=" <No file selected>", font=FONT_MAIN, bg=WIN95_WHITE, relief=tk.SUNKEN, bd=2, anchor="w", width=40)
        self.lbl_ytd.grid(row=1, column=1, padx=10, pady=5)
        tk.Button(group1, text="Browse...", font=FONT_MAIN, bg=WIN95_BG, relief=tk.RAISED, bd=2, width=10, command=self.select_ytd_file).grid(row=1, column=2)

        # 2. Configuration
        group2 = tk.LabelFrame(main_frame, text=" 2. Configuration ", font=FONT_MAIN, bg=WIN95_BG, relief=tk.GROOVE, bd=2, padx=10, pady=10)
        group2.pack(fill="x", pady=(0, 10))

        chk_header = tk.Frame(group2, bg=WIN95_BG)
        chk_header.pack(fill="x", anchor="w")
        tk.Label(chk_header, text="Select sheets to audit:", font=FONT_MAIN, bg=WIN95_BG).pack(side="left")
        tk.Button(chk_header, text="Select All / None", font=FONT_MAIN, bg=WIN95_BG, relief=tk.RAISED, bd=1, command=self.toggle_select_all).pack(side="right")
        
        list_container = tk.Frame(group2, bg=WIN95_WHITE, relief=tk.SUNKEN, bd=2)
        list_container.pack(fill="x", pady=5)
        
        self.canvas = tk.Canvas(list_container, bg=WIN95_WHITE, height=80, highlightthickness=0)
        self.scrollbar = tk.Scrollbar(list_container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg=WIN95_WHITE)

        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        self.root.bind_all("<MouseWheel>", self._on_mousewheel)

        out_frame = tk.Frame(group2, bg=WIN95_BG)
        out_frame.pack(fill="x", pady=(10, 0))
        tk.Label(out_frame, text="Save To:", font=FONT_MAIN, bg=WIN95_BG).pack(side="left")
        self.lbl_output = tk.Label(out_frame, text=" <No location>", font=FONT_MAIN, bg=WIN95_WHITE, relief=tk.SUNKEN, bd=2, anchor="w", width=36)
        self.lbl_output.pack(side="left", padx=10)
        tk.Button(out_frame, text="Save As...", font=FONT_MAIN, bg=WIN95_BG, relief=tk.RAISED, bd=2, width=10, command=self.select_output_file).pack(side="left")

        # 3. Execution
        group3 = tk.LabelFrame(main_frame, text=" 3. Execution ", font=FONT_MAIN, bg=WIN95_BG, relief=tk.GROOVE, bd=2, padx=10, pady=10)
        group3.pack(fill="both", expand=True)

        self.progress_canvas = tk.Canvas(group3, height=18, bg=WIN95_WHITE, relief=tk.SUNKEN, bd=2)
        self.progress_canvas.pack(fill="x", pady=(5, 10))

        btn_frame = tk.Frame(group3, bg=WIN95_BG)
        btn_frame.pack(fill="x")
        self.btn_run = tk.Button(btn_frame, text="Execute Audit", font=FONT_BOLD, bg=WIN95_BG, relief=tk.RAISED, bd=2, width=15, pady=3, command=self.start_thread)
        self.btn_run.pack(side="right")
        tk.Button(btn_frame, text="Exit", font=FONT_MAIN, bg=WIN95_BG, relief=tk.RAISED, bd=2, width=10, pady=3, command=self.root.destroy).pack(side="right", padx=10)

        self.status_bar = tk.Label(self.root, text=" System Initialized. Awaiting input parameters...", font=FONT_MAIN, bg=WIN95_BG, relief=tk.SUNKEN, bd=1, anchor="w")
        self.status_bar.pack(side="bottom", fill="x")

    def log(self, message):
        self.status_bar.config(text=f" {message}")
        self.root.update_idletasks()

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    def toggle_select_all(self):
        all_selected = all(var.get() for var in self.sheet_vars.values())
        new_state = not all_selected
        for var in self.sheet_vars.values():
            var.set(new_state)

    def update_progress(self, current, total):
        self.progress_canvas.delete("all")
        if total > 0:
            percentage = current / total
            width = self.progress_canvas.winfo_width()
            fill_width = int(width * percentage)
            block_size = 12
            for x in range(2, fill_width, block_size + 2):
                self.progress_canvas.create_rectangle(x, 2, x + block_size, 16, fill=WIN95_BLUE, outline=WIN95_BLUE)
        self.root.update_idletasks()

    def select_monthly_files(self):
        files = filedialog.askopenfilenames(title="Select Monthly Excel Files", filetypes=[("Excel files", "*.xlsx *.xls")])
        if files:
            self.monthly_files = list(files)
            self.lbl_monthly.config(text=f" {len(self.monthly_files)} file(s) loaded")
            self.log(f"Loaded {len(self.monthly_files)} monthly files.")

    def select_ytd_file(self):
        file = filedialog.askopenfilename(title="Select YTD Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])
        if file:
            self.ytd_file = file
            self.lbl_ytd.config(text=f" {os.path.basename(file)}")
            self.log(f"Loaded YTD file: {os.path.basename(file)}")
            self.load_sheets()

    def select_output_file(self):
        file = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="Audit_Result.xlsx", title="Save Result As", filetypes=[("Excel files", "*.xlsx")])
        if file:
            self.output_file = file
            self.lbl_output.config(text=f" {os.path.basename(file)}")

    def load_sheets(self):
        self.log("Reading workbook structure...")
        try:
            xl = pd.ExcelFile(self.ytd_file)
            for widget in self.scrollable_frame.winfo_children():
                widget.destroy()
            self.sheet_vars.clear()

            ignored = ['help', 'instruction', 'master']
            for sheet in xl.sheet_names:
                if not any(word in sheet.lower() for word in ignored):
                    var = tk.BooleanVar(value=False) 
                    self.sheet_vars[sheet] = var
                    cb = tk.Checkbutton(self.scrollable_frame, text=sheet, variable=var, bg=WIN95_WHITE, activebackground=WIN95_WHITE, font=FONT_MAIN)
                    cb.pack(anchor="w", padx=5)
            self.log("Ready. Select target sheets.")
        except Exception as e:
            self.log(f"Error: {e}")

    # ==========================================
    # UPDATED SMART KEY LOGIC FOR PARTY CHANGES
    # ==========================================
    def get_primary_keys(self, columns):
        cols = [str(c).strip() for c in columns]
        keys = []
        
        # Priority 1: Absolute Identifiers (Note No, Invoice No)
        doc_ids = ['Invoice Number', 'Revised Invoice Number', 'Note Number', 'Revised Note Number', 'Document Number', 'Revised Document Number', 'Shipping Bill Number']
        for col in cols:
            if col in doc_ids: keys.append(col)
            
        # IMPORTANT: If we found a Note Number or Invoice Number, WE STOP HERE.
        # This forces the system to treat GSTIN/Name changes as modifications to the existing row!
        if keys:
            return keys
            
        # Priority 2: Composites (for b2cs and similar sheets without doc numbers)
        composites = ['Place Of Supply', 'Original Place Of Supply', 'Rate', 'Type', 'Financial Year', 'Original Month', 'Nature of Supply', 'Export Type', 'HSN', 'Nature of Document', 'Description']
        for col in cols:
            if col in composites: keys.append(col)
            
        # Priority 3: Context Identifiers
        context_ids = ['GSTIN/UIN of Recipient', 'Supplier GSTIN/UIN', 'Recipient GSTIN/UIN', 'E-Commerce GSTIN', 'UQC', 'Sr. No. From']
        for col in cols:
            if col in context_ids and col not in keys: keys.append(col)
            
        return keys

    def start_thread(self):
        threading.Thread(target=self.run_comparison, daemon=True).start()

    def run_comparison(self):
        selected_sheets = [sheet for sheet, var in self.sheet_vars.items() if var.get()]
        
        if not self.monthly_files or not self.ytd_file or not self.output_file:
            messagebox.showwarning("Warning", "Configuration incomplete.", icon='warning')
            return
        if not selected_sheets:
            messagebox.showwarning("Warning", "No sheets selected.", icon='warning')
            return
            
        self.btn_run.config(state='disabled', relief=tk.SUNKEN)
        self.update_progress(0, len(selected_sheets))
        self.log("Executing Two-Way Audit Protocol... Please wait.")

        system_log = []

        try:
            try:
                with open(self.output_file, 'a'): pass
            except IOError:
                messagebox.showerror("I/O Error", "Output file is open in Excel. Please close it first.")
                return

            sheets_with_data = 0
            
            with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
                
                for idx, sheet in enumerate(selected_sheets):
                    try:
                        self.log(f"Analyzing: {sheet}...")
                        
                        df_ytd = pd.read_excel(self.ytd_file, sheet_name=sheet, header=3)
                        df_ytd = df_ytd.loc[:, ~df_ytd.columns.duplicated()]
                        df_ytd = df_ytd.loc[:, ~df_ytd.columns.astype(str).str.contains('^Unnamed', na=False)]
                        df_ytd.columns = [str(c).strip() for c in df_ytd.columns]
                        
                        keys = self.get_primary_keys(df_ytd.columns)
                        if not keys:
                            system_log.append({"Sheet Name": sheet, "Audit Status": "SKIPPED", "Differences Found": 0, "Details": "Could not identify Primary Keys (Invoice Number, etc)."})
                            continue
                            
                        monthly_dfs = []
                        for m_file in self.monthly_files:
                            try:
                                xl = pd.ExcelFile(m_file)
                                if sheet in xl.sheet_names:
                                    m_df = pd.read_excel(m_file, sheet_name=sheet, header=3)
                                    m_df = m_df.loc[:, ~m_df.columns.duplicated()]
                                    m_df = m_df.loc[:, ~m_df.columns.astype(str).str.contains('^Unnamed', na=False)]
                                    m_df.columns = [str(c).strip() for c in m_df.columns]
                                    monthly_dfs.append(m_df)
                            except: pass
                            
                        if not monthly_dfs:
                            df_ytd['Audit Status'] = 'New Backdated Entry'
                            df_ytd.to_excel(writer, sheet_name=sheet[:31], index=False)
                            sheets_with_data += 1
                            system_log.append({"Sheet Name": sheet, "Audit Status": "WARNING", "Differences Found": len(df_ytd), "Details": "No historical data found. All entries marked as New."})
                        else:
                            df_snapshots = pd.concat(monthly_dfs, ignore_index=True)
                            
                            df_ytd.dropna(subset=keys, how='all', inplace=True)
                            snap_keys = [k for k in keys if k in df_snapshots.columns]
                            if snap_keys:
                                df_snapshots.dropna(subset=snap_keys, how='all', inplace=True)
                                
                            for k in keys:
                                df_ytd[k] = df_ytd[k].astype(str).str.strip()
                                if k in df_snapshots.columns:
                                    df_snapshots[k] = df_snapshots[k].astype(str).str.strip()
                                else:
                                    df_snapshots[k] = ""
                                    
                            df_snapshots.drop_duplicates(subset=keys, keep='last', inplace=True)
                            
                            cols_to_compare = [c for c in df_ytd.columns if c not in keys]
                            for c in cols_to_compare:
                                if c not in df_snapshots.columns:
                                    df_snapshots[c] = np.nan
                                    
                            merged = pd.merge(df_ytd, df_snapshots[keys + cols_to_compare], on=keys, how='outer', suffixes=('', '_old'), indicator=True)
                            
                            if merged.empty:
                                system_log.append({"Sheet Name": sheet, "Audit Status": "SUCCESS", "Differences Found": 0, "Details": "Empty DataFrames merged."})
                                continue
                                
                            def check_row(row):
                                if row['_merge'] == 'left_only': return 'New Backdated Entry'
                                if row['_merge'] == 'right_only': return 'Deleted / Missing from YTD'
                                
                                changes = []
                                for col in cols_to_compare:
                                    try:
                                        new_val, old_val = row[col], row[f"{col}_old"]
                                        if pd.isna(new_val) and pd.isna(old_val): continue
                                        
                                        nv_str = "" if pd.isna(new_val) else str(new_val).strip()
                                        ov_str = "" if pd.isna(old_val) else str(old_val).strip()
                                        if nv_str.lower() == 'nan': nv_str = ""
                                        if ov_str.lower() == 'nan': ov_str = ""
                                        if nv_str == ov_str: continue
                                        
                                        if 'date' in col.lower():
                                            try:
                                                nd, od = pd.to_datetime(new_val, errors='coerce'), pd.to_datetime(old_val, errors='coerce')
                                                if pd.notna(nd) and pd.notna(od):
                                                    if nd.date() != od.date():
                                                        changes.append(f"{col}: {od.strftime('%d-%b-%Y')} -> {nd.strftime('%d-%b-%Y')}")
                                                    continue 
                                            except: pass

                                        try:
                                            if round(float(new_val), 2) != round(float(old_val), 2):
                                                changes.append(f"{col}: {ov_str} -> {nv_str}")
                                            continue 
                                        except: pass
                                            
                                        if nv_str.upper() != ov_str.upper():
                                            changes.append(f"{col}: {ov_str} -> {nv_str}")
                                    except: pass 
                                        
                                if changes: return 'Modified: ' + ' | '.join(changes)
                                return 'No Change'

                            merged['Audit Status'] = merged.apply(check_row, axis=1)
                            
                            for col in cols_to_compare:
                                mask = merged['_merge'] == 'right_only'
                                try:
                                    merged.loc[mask, col] = merged.loc[mask, f"{col}_old"]
                                except: pass

                            action_df = merged[merged['Audit Status'] != 'No Change'].copy()
                            action_df.drop(columns=[f"{c}_old" for c in cols_to_compare] + ['_merge'], inplace=True, errors='ignore')
                            
                            if not action_df.empty:
                                action_df.to_excel(writer, sheet_name=sheet[:31], index=False)
                                sheets_with_data += 1
                                system_log.append({"Sheet Name": sheet, "Audit Status": "DIFFERENCES FOUND", "Differences Found": len(action_df), "Details": "Processed cleanly."})
                                
                                worksheet = writer.sheets[sheet[:31]]
                                
                                header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
                                odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                                even_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid") 
                                
                                for cell in worksheet[1]:
                                    cell.fill = header_fill
                                    cell.font = Font(bold=True, color="FFFFFF")
                                    cell.alignment = Alignment(horizontal="center", vertical="center")
                                    
                                for col in worksheet.columns:
                                    max_length = 0
                                    col_letter = col[0].column_letter
                                    for cell in col:
                                        try:
                                            if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                                        except: pass
                                    worksheet.column_dimensions[col_letter].width = min(max_length + 3, 50) 
                                    
                                status_col_idx = len(action_df.columns)
                                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                                    current_fill = even_fill if row_idx % 2 == 0 else odd_fill
                                    for cell in row:
                                        cell.fill = current_fill
                                        if cell.column == status_col_idx:
                                            if "New" in str(cell.value):
                                                cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid") 
                                                cell.font = Font(color="155724", bold=True)
                                            elif "Modified" in str(cell.value):
                                                cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid") 
                                                cell.font = Font(color="856404", bold=True)
                                            elif "Deleted" in str(cell.value):
                                                cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid") 
                                                cell.font = Font(color="721C24", bold=True)
                                worksheet.freeze_panes = 'A2'
                            else:
                                system_log.append({"Sheet Name": sheet, "Audit Status": "SUCCESS", "Differences Found": 0, "Details": "Data matches perfectly."})
                                
                    except Exception as sheet_err:
                        error_details = traceback.format_exc().splitlines()[-1] 
                        system_log.append({"Sheet Name": sheet, "Audit Status": "CRASHED", "Differences Found": "ERROR", "Details": str(error_details)})
                    finally:
                        self.update_progress(idx + 1, len(selected_sheets))

                log_df = pd.DataFrame(system_log)
                log_df.to_excel(writer, sheet_name="System_Log", index=False)
                worksheet = writer.sheets["System_Log"]
                
                header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = Font(bold=True, color="FFFFFF")
                
                worksheet.column_dimensions['A'].width = 20
                worksheet.column_dimensions['B'].width = 20
                worksheet.column_dimensions['C'].width = 20
                worksheet.column_dimensions['D'].width = 80
                
                for row in worksheet.iter_rows(min_row=2):
                    if "CRASHED" in str(row[1].value):
                        for cell in row: cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    elif "SUCCESS" in str(row[1].value):
                        for cell in row: cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    elif "DIFFERENCES" in str(row[1].value):
                        for cell in row: cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

            self.log("Operation Complete. Check System_Log tab in output file.")
            try:
                os.startfile(os.path.abspath(self.output_file))
            except: pass

        except Exception as e:
            self.log("Operation Aborted due to critical system error.")
            messagebox.showerror("System Error", f"Operation aborted:\n{str(e)}", icon='error')
        finally:
            self.btn_run.config(state='normal', relief=tk.RAISED)

if __name__ == "__main__":
    root = tk.Tk()
    app = RetroGSTApp(root)
    root.mainloop()
